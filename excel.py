import openpyxl as xl


wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1,1)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9